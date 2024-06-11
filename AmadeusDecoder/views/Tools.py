'''
Created on 8 Sep 2022

'''
import base64
import datetime
import io
from itertools import count
import json
import os
import shlex
import subprocess
import time

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt

import requests

from AmadeusDecoder.models.invoice.InvoicePassenger import PassengerInvoice
from AmadeusDecoder.models.invoice.Ticket import Ticket
from AmadeusDecoder.models.pnr.Passenger import Passenger
from AmadeusDecoder.models.pnr.Pnr import Pnr
from AmadeusDecoder.models.invoice.Ticket import Ticket
from AmadeusDecoder.models.invoice.Fee import Fee, OthersFee

from AmadeusDecoder.models.pnrelements.Airline import Airline
from AmadeusDecoder.models.pnrelements.Airport import Airport
from AmadeusDecoder.models.pnrelements.PnrAirSegments import PnrAirSegments
from AmadeusDecoder.models.user.Users import User
from AmadeusDecoder.utilities.ProductImportParser import ProductParser, CustomerParser
from AmadeusDecoder.models.configuration.Configuration import Configuration
from AmadeusDecoder.models.invoice.CancelSaleOrder import CancelSaleOrder

from AmadeusDecoder.models.utilities.Comments import Comment

from django.db import connection
from django.db.models import Count, Func
from django.db.models.functions import Lower, ExtractYear, Cast
from openpyxl import Workbook
from openpyxl.styles import *
import decimal
from datetime import datetime, timedelta

@csrf_exempt
def cancel_order_sale_from_odoo(request):
    if request.method == 'POST':
        pnr_number = request.POST.get('pnr_number')
        invoice_number = request.POST.get('invoice_number')
        agent = request.POST.get('agent')
        # client = request.POST.get('client')
        
        print(pnr_number, invoice_number, agent)
        
        if pnr_number is not None and invoice_number is not None:
            try:
                pnr = Pnr.objects.get(number=pnr_number)
                passenger_invoices = PassengerInvoice.objects.filter(pnr_id=pnr.id, invoice_number=invoice_number).all()
                
                if passenger_invoices:
                    for passenger_invoice in passenger_invoices:
                        # Supprimer la facture passager correspondante si elle existe
                        PassengerInvoice.objects.filter(id=passenger_invoice.id).delete()
                        
                        if passenger_invoice.ticket_id:
                            # Supprimer le billet correspondant s'il existe
                            Ticket.objects.filter(id=passenger_invoice.ticket_id).update(is_invoiced=False)
                        
                        if passenger_invoice.fee_id:
                            # Supprimer les frais correspondants s'ils existent
                            Fee.objects.filter(id=passenger_invoice.fee_id).update(is_invoiced=False)
                            
                        if passenger_invoice.other_fee_id:
                            # Supprimer les autres frais correspondants s'ils existent
                            OthersFee.objects.filter(id=passenger_invoice.other_fee_id).update(is_invoiced=False)
                            
                    
                        print("HELLO")
                        print(passenger_invoice.client, passenger_invoice.ticket, passenger_invoice.fee)    
                        
                        cancel_sale_order = CancelSaleOrder(
                                                pnr=pnr,
                                                client=passenger_invoice.client,
                                                agent=agent,
                                                ticket=passenger_invoice.ticket,
                                                fee=passenger_invoice.fee,
                                                other_fee=passenger_invoice.other_fee,
                                                invoice_number=invoice_number
                                            )
                        
                        cancel_sale_order.save()                            
                    
                    return JsonResponse({'status': 'ok', 'message': f'PNR {pnr_number} a été décommandé sur gestion PNR'}, status=200)
            except Pnr.DoesNotExist:
                return JsonResponse({'message': 'Numéro de PNR introuvable'}, status=404)
            except PassengerInvoice.DoesNotExist:
                return JsonResponse({'message': 'Numéro de facture introuvable'}, status=404)
        else:
            return JsonResponse({'message': 'Numéro de PNR ou de facture non fourni'}, status=400)
    else:
        return JsonResponse({'message': 'Seules les requêtes POST sont autorisées'}, status=405)

@login_required(login_url='index')
def tools(request):  
    return render(request,'tools.html')


def call_product_import(request):
    
    status = ''
    try:
        directory = '/opt/odoo/issoufali-addons/sync_products/data/exported/'
        for file in os.listdir(directory):
            if os.path.isfile(os.path.join(directory, file)):
                ProductParser.import_product(os.path.join(directory, file), directory)
        status = 'Product successfuly imported'
    except Exception as e:
        print(e)
        status = str(e)

    return JsonResponse({'return':status})


def call_customer_import(request):
    
    status = ''
    try:
        directory = '/opt/odoo/issoufali-addons/export_contacts/data/exported/'
        for file in os.listdir(directory):
            if os.path.isfile(os.path.join(directory, file)):
                CustomerParser.import_customer(os.path.join(directory, file), directory)
        status = 'Customer successfuly imported'
    except Exception as e:
        print(e)
        status = str(e)


    return JsonResponse({'return':status})

    #get the invoice number for unordering a pnr
def get_invoice_number(request,numeroPnr):
    
    pnr = Pnr.objects.get(number=numeroPnr)
    invoices = PassengerInvoice.objects.filter(pnr_id=pnr.id, is_invoiced=True).distinct()

    invoices_data = [{'invoice_number': invoice.invoice_number} for invoice in invoices]
    
    unique_invoice_numbers_set = {entry['invoice_number'] for entry in invoices_data}

    unique_invoice_numbers_list = list(unique_invoice_numbers_set)

    print(unique_invoice_numbers_list)
    return JsonResponse({'invoices': unique_invoice_numbers_list})


# ------------------------ EXPORT EXCEL ----------------------------------------------------------------


def get_route_by_ticket(ticket_id):
    ticket = Ticket.objects.get(pk=ticket_id)
    route = ''
    for passengerSegment in ticket.ticket_parts.all().order_by('segment__id'):
        if passengerSegment.segment.segment_type == 'SVC':
            route = 'SVC'
        else:
            if passengerSegment.segment.codeorg is not None:
                route += passengerSegment.segment.codeorg.iata_code + '/' + passengerSegment.segment.codedest.iata_code
            route += '//'
    if route.endswith('//'):
        route = route.removesuffix('//')
    return route

def get_segment_by_ticket(ticket_id):
    ticket = Ticket.objects.get(pk=ticket_id)
    segments = ''
    for passengerSegment in ticket.ticket_parts.all().order_by('segment__id'):
        segments += passengerSegment.segment.segmentorder + '-'
    for ssrs in  ticket.ticket_ssrs.all():
        segments += ssrs.ssr.order_line + '-'
    return segments[:-1]

def get_tst_passengers(tst_id):
    tst = Ticket.objects.filter(id=tst_id).first()
    passengers = ''
    for ticket_passenger_tst in tst.ticket_tst_parts.all():
        passengers += ticket_passenger_tst.passenger.order + '-'
    return passengers[:-1]

def pnr_to_excel(request,pnr_id):
    
    pnr = Pnr.objects.get(pk=pnr_id)


    tickets = Ticket.objects.filter(pnr_id=pnr_id).all()
    air_segments = pnr.segments.filter(segment_type='Flight', air_segment_status=1).all().order_by('segmentorder')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'PNR details' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:K1')
    worksheet.merge_cells('A2:K2')

    first_cell = worksheet['A1']
    first_cell.value = "Details du PNR " + pnr.number
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    second_cell = worksheet['A2']
    second_cell.value = pnr.number
    second_cell.font = Font(bold=True, color="246ba1")
    second_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Details du PNR' + " " + pnr.number
    
    # -------------------------- SEGMENT SECTION --------------------------------------
    worksheet.merge_cells('A4:I4')

    first_cell = worksheet['A4']
    first_cell.value = "Informations de vol "
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    segment_columns = ['Segment','Vols','Classe','Cabine','Départ','Arrivée','Date et heure de départ',"Date et heure d'arrivée",'OPC']
    segment_row_num = 6

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(segment_columns, 1):
        print("col_num : ", col_num)
        print('column_title : ', column_title)
        cell = worksheet.cell(row=segment_row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid",fgColor="50C878")
        cell.font = Font(bold=True, color="F7F6FA")
        third_cell = worksheet['K4']
        third_cell.alignment = Alignment(horizontal="right")
    segment_row_num += 1    

    for segment in air_segments :
        segment_row_num += 1

         # Define data for each cell in the row
        segment_row = [segment.segmentorder ,segment.__str__(),segment.flightclass ,get_segment_cabin(segment),segment.codeorg.iata_code,segment.codedest.iata_code,'','','']

        if segment.segment_state == 0 :
            segment_row[6] = segment.departuretime.strftime('%d/%m/%Y %H:%M')
        else:
            segment_row[6] = segment.departuretime.strftime('%d/%m/%Y')

        if segment.segment_state == 0 :
            segment_row[7] = segment.arrivaltime.strftime('%d/%m/%Y %H:%M')
        else:
            segment_row[7] = 'Flown'

        if pnr.status != 'Emis':
            if get_opc(segment) is not None:
                segment_row[8] = get_opc(segment).strftime('%d/%m/%Y %H:%M')
            
        
        # Assign data for each cell of the segment row
        for col_num, cell_value in enumerate(segment_row, 1):
            cell = worksheet.cell(row=segment_row_num, column = col_num)
            cell.value = cell_value
            if isinstance(cell_value, decimal.Decimal):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1


    #  ----------------------- TICKET SECTION --------------------------------------
    # Define columns title
    columns = ['Type','Article','Passager(s)/Trajet','Transport','Taxe','Total','Passager/Segment(s)',"Date d'émission",'Date de commande','Numéro de commade','Emetteur']
    ticket_row_num = segment_row_num +2
    fee_row_num = ticket_row_num +1
    invoice_number = 0

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=ticket_row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid",fgColor="50C878")
        cell.font = Font(bold=True, color="F7F6FA")
        third_cell = worksheet['K3']
        third_cell.alignment = Alignment(horizontal="right")
    ticket_row_num += 1    
    
    for ticket in tickets:
        route = get_route_by_ticket(ticket.id)

        if ticket.emitter_id is not None:
            emitter = User.objects.get(pk=ticket.emitter_id).username
        else:
            emitter = " "
            
        if ticket.issuing_date is not None:
            ticket_issuing_date = ticket.issuing_date.strftime("%d/%m/%y")
        else:
            ticket_issuing_date= ''

        # Define data for each cell in the row
        ticket_row = [ticket.ticket_type,ticket.number,'',ticket.transport_cost,ticket.tax,ticket.total,'',ticket_issuing_date,'','',emitter]

        # Assign data to the Passenger/Trajet column
        if ticket.ticket_type == 'EMD' and pnr.type == 'EWA':
            if ticket.ticket_description is not None:
                ticket_row[2] = ticket.ticket_description

        if ticket.passenger is not None:
            ticket_row[2] = ticket.passenger.__str__() +" / "+route
        else:
            ticket_row[2] = get_tst_passengers(ticket.id) +" / "+route

        # Assign data to the Passenger/segment(s) column
        if ticket.passenger.order is not None:
            ticket_row[6] = ticket.passenger.order + " / " + get_segment_by_ticket(ticket.id)
        else:
            ticket_row[6] = get_tst_passengers(ticket.id) + " / " + get_segment_by_ticket(ticket.id)

        # get invoice details
        passenger_invoices = PassengerInvoice.objects.filter(pnr_id=pnr_id, ticket_id=ticket.id, is_invoiced=True)
        if passenger_invoices.exists():
            for passenger_invoice in passenger_invoices:
                invoice_number = passenger_invoice.invoice_number
                invoice_date = passenger_invoice.date_creation.date()

                ticket_row[8] = invoice_date.strftime("%d/%m/%y")  # Update invoice_date field
                ticket_row[9] = invoice_number # Update invoice_number field
        
        fees = Fee.objects.filter(pnr_id=pnr_id, ticket_id=ticket.id)
        if fees.exists():
            for fee in fees:
                
                fee_row = [
                "Fee",  # Ticket type
                fee.type,  # Ticket number
                "",  # Passenger name (empty for fees)
                fee.cost,  # Transport cost
                fee.tax,  # Tax
                fee.total,  # Total
                "",  # Placeholder for empty field
                ticket.issuing_date.strftime("%d/%m/%y"),  # Issuing date
                "",  # Placeholder for empty field
                "",  # Placeholder for empty field
                emitter  # Emitter
            ]
            
            if passenger_invoices.exists():
                fee_row[8] = invoice_date.strftime("%d/%m/%y") # Invoice de=ate
                fee_row[9] = invoice_number  # invoice number
                
            fee_row_num = ticket_row_num + 1
                

        # Assign data for each cell of the ticket row
        for col_num, cell_value in enumerate(ticket_row, 1):
            cell = worksheet.cell(row=ticket_row_num, column = col_num)
            cell.value = cell_value
            if isinstance(cell_value, decimal.Decimal):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        # Assign data for each cell of the fee row
        for col_num, cell_value in enumerate(fee_row, 1):
            cell = worksheet.cell(row=fee_row_num, column = col_num)
            cell.value = cell_value
            if isinstance(cell_value, decimal.Decimal):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        ticket_row_num += 2
    
        
    workbook.save(response)
    return response

# get flight cabin from segment's flight class
def get_segment_cabin(segment):
    from AmadeusDecoder.models.invoice.ServiceFees import ClassCabin
    
    flight_cabin = ''
    try:
        related_pnr = segment.pnr
        related_cabin = ClassCabin.objects.filter(sign__contains=[segment.flightclass], gdsprovider=related_pnr.type).first()
        if related_cabin is not None:
            return related_cabin.type
    except Exception as e:
        return flight_cabin
    
# get opc for the segment informations
def get_opc(segment, ssr):
    from AmadeusDecoder.models.pnrelements.ConfirmationDeadline import ConfirmationDeadline
    confirmation_deadline = ConfirmationDeadline()
    if segment is not None:
        confirmation_deadline.segment = segment
    elif ssr is not None:
        confirmation_deadline.ssr = ssr
        
    try:
        return confirmation_deadline.get_confirmation_deadline().doc_date
    except:
        return ''


def pnr_list_to_excel(request):
    context = {}
    if request.method == 'POST':
        pnr_list_id = json.loads(request.POST.get('pnr_list'))
        print(pnr_list_id)
        pnr_list = []
        for pnr_id in pnr_list_id:
            pnr = Pnr.objects.get(pk=pnr_id)
            pnr_list.append(pnr)

        all_pnr = []

        # Define data for each cell of the row
        for pnr in pnr_list:
            pnr_row = {}
            pnr_row['pnr_number'] = pnr.number
            # get the first passenger name
            first_passenger = get_first_passenger(pnr)
            if first_passenger is not None:
                if first_passenger.name is not None and first_passenger.surname is not None:
                    pnr_row['passenger'] = first_passenger.name + ' '+ first_passenger.surname
                else:
                    pnr_row['passenger'] = first_passenger.surname
            else:
                pnr_row['passenger'] = ''

            # get the customer
            customer = get_customer_in_passenger_invoice(pnr.id)
            if pnr.status == 'Non émis':
                pnr_row['customer'] = pnr.customer.intitule
            else:
                if customer is not None:
                    pnr_row['customer'] = customer.intitule
                else:
                    pnr_row['customer'] = ''

            # creation date
            pnr_row['creation_date'] = pnr.system_creation_date.strftime('%d/%m/%Y %H:%M')

            # ticket issuing date
            ticket_issuing_date = ticket_get_issuing_date(pnr)
            if ticket_issuing_date is not None and pnr.status_value == 0:
                pnr_row['ticket_issuing_date'] = ticket_issuing_date.strftime('%d/%m/%Y')
            else:
                pnr_row['ticket_issuing_date'] = ''

            # total amount
            pnr_row['montant'] = round(get_order_amout_total(pnr),2 ) 
            
            # pnr status
            pnr_row['pnr_status'] = pnr.status  

            # OPC
            opc = get_min_opc(pnr)
            if opc is not None:
                pnr_row['opc'] = opc.strftime('%d/%m/%Y %H:%M')
            else:
                pnr_row['opc'] = ''

            # pnr type
            if pnr.type == 'EWA':
                pnr_row['type'] = 'Zenith'
            else:
                pnr_row['type'] = pnr.type
            
            # pnr creator agent
            pnr_creator = pnr.get_creator_agent()
            if pnr_creator is not None:
                if not isinstance(pnr_creator, str):
                    pnr_row['creator_agent'] = pnr_creator.username
                else:
                    pnr_row['creator_agent'] = pnr_creator
            else:
                pnr_row['creator_agent'] = ''

            # pnr emitter agent
            pnr_emitter = pnr.get_emit_agent()
            if pnr_emitter is not None:
                if not isinstance(pnr_emitter, str):
                    pnr_row['emitter_agent'] = pnr_emitter.username
                else:
                    pnr_row['emitter_agent'] = pnr_emitter
            else:
                pnr_row['emitter_agent'] = ''
                
            # pnr office
            pnr_office = get_pnr_office(pnr)
            if pnr_office is not None:
                pnr_row['pnr_office'] = pnr_office
            else:
                pnr_row['pnr_office'] = ''

            # pnr agency
            if pnr.agency is not None:
                pnr_row['pnr_agency'] = pnr.agency.code
            else:
                pnr_row['pnr_agency'] = ''

            all_pnr.append(pnr_row)
        context['results'] = all_pnr
        return JsonResponse(context)


# ------------------------- DETAILS PNR -----------------------------------------------------------
def get_first_passenger(pnr):
    from AmadeusDecoder.models.pnr.Passenger import Passenger
    try:
        passenger = Passenger.objects.filter(passenger__pnr=pnr).first()
        if passenger is not None:
            return passenger
        else:
            return None
    except:
        return None  
        
def get_customer_in_passenger_invoice(pnr_id):
    """_summary_
    Returns:
        PNR issued : get all customers in passenger_invoice filtered by pnr_id
        PNR not issued : get customer directly on PNR (customer_id)
    """
    from AmadeusDecoder.models.invoice.InvoicePassenger import PassengerInvoice
    from AmadeusDecoder.models.invoice.Clients import Client
    
    if pnr_id != '' and pnr_id != None:                
        passenger_invoice_obj = PassengerInvoice.objects.filter(pnr_id=pnr_id)
        if passenger_invoice_obj.exists():            
            for passenger_invoice in passenger_invoice_obj:
                client_obj = Client.objects.filter(id=passenger_invoice.client_id)
                    
            if client_obj.exists():
                for client in client_obj:
                    return client
        else:
            return None
    return None

def ticket_get_issuing_date(pnr):
    try:
        return pnr.get_max_issuing_date()
    except:
        return None
    
def get_order_amout_total(pnr):
    from AmadeusDecoder.models.invoice.InvoicePassenger import PassengerInvoice
    from AmadeusDecoder.models.invoice.Ticket import Ticket
    from AmadeusDecoder.models.invoice.Fee import OthersFee, Fee
    pnr = Pnr.objects.get(pk=pnr.id)
    passenger_invoices = PassengerInvoice.objects.filter(pnr_id=pnr)
    amount_total = 0
    amount_invoiced = 0
    pnr_elements_count = 0
    fee_not_invoiced_count = 0
    tickets = Ticket.objects.filter(pnr=pnr, ticket_status=1, is_invoiced=False)
    other_fees = OthersFee.objects.filter(pnr=pnr, is_invoiced=False)
    if tickets.exists() and other_fees.exists():
        for ticket in tickets:
            fees = Fee.objects.filter(ticket=ticket, is_invoiced=False)
            if fees.exists():
                fee_not_invoiced_count += 1
        for other_fee in other_fees:
            fees = Fee.objects.filter(other_fee=other_fee, is_invoiced=False)
            if fees.exists():
                fee_not_invoiced_count += 1
        pnr_elements_count = tickets.count() + other_fees.count() + fee_not_invoiced_count
    if pnr.status_value == 0:
        order_invoiced = passenger_invoices.filter(status='sale', is_invoiced=True)
        passenger_invoice = passenger_invoices.filter(status='sale', is_invoiced=False)
        if order_invoiced.exists():
            for order in order_invoiced:
                if order.ticket is not None and order.ticket.ticket_status == 1:
                    amount_invoiced += order.ticket.total
                if order.fee is not None:
                    if (order.fee.ticket is not None and order.fee.ticket.ticket_status == 1) or (order.fee.other_fee is not None and order.fee.other_fee.other_fee_status == 1):
                        amount_invoiced += order.fee.total
                if order.other_fee is not None and order.other_fee.other_fee_status == 1:
                    amount_invoiced += order.other_fee.total
            if passenger_invoice.exists() and passenger_invoice.count() == pnr_elements_count:
                for order in passenger_invoice:
                    if order.ticket is not None and order.ticket.ticket_status == 1:
                        amount_total += order.ticket.total
                    if order.fee is not None:
                        if (order.fee.ticket is not None and order.fee.ticket.ticket_status == 1) or (order.fee.other_fee is not None and order.fee.other_fee.other_fee_status == 1):
                            amount_invoiced += order.fee.total
                    if order.other_fee is not None and order.other_fee.other_fee_status == 1:
                        amount_total += order.other_fee.total
            else:
                amount_total = pnr.invoice.detail.total - amount_invoiced
        else:
            amount_total = pnr.invoice.detail.total
    elif pnr.status_value == 1:
        quotation_invoiced = passenger_invoices.filter(status='quotation', is_quotation=True)
        passenger_invoice = passenger_invoices.filter(status='quotation', is_quotation=False)
        if quotation_invoiced.exists():
            for order in quotation_invoiced:
                if order.ticket is not None and order.ticket.ticket_status == 1:
                    amount_invoiced += order.ticket.total
                if order.fee is not None:
                    if (order.fee.ticket is not None and order.fee.ticket.ticket_status == 1) or (order.fee.other_fee is not None and order.fee.other_fee.other_fee_status == 1):
                        amount_invoiced += order.fee.total
                if order.other_fee is not None and order.other_fee.other_fee_status == 1:
                    amount_invoiced += order.other_fee.total
                if order.invoice_id is not None:
                    amount_total += order.invoice_id.detail.total
            if passenger_invoice.exists() and passenger_invoice.count() == pnr_elements_count:
                for order in passenger_invoice:
                    if order.ticket is not None and order.ticket.ticket_status == 1:
                        amount_total += order.ticket.total
                    if order.fee is not None:
                        if (order.fee.ticket is not None and order.fee.ticket.ticket_status == 1) or (order.fee.other_fee is not None and order.fee.other_fee.other_fee_status == 1):
                            amount_invoiced += order.fee.total
                    if order.other_fee is not None and order.other_fee.other_fee_status == 1:
                        amount_total += order.other_fee.total
                    if order.invoice_id is not None:
                        amount_total += order.invoice_id.detail.total
        else:
            amount_total = pnr.invoice.detail.total

    return amount_total

def get_min_opc(pnr):
    try:
        return pnr.get_min_opc()
    except:
        return None
    
def get_pnr_office(pnr):
    try:   
        # Make agency name uniformised     
        agence_name_uniformised = ['GSA ISSOUFALI Dzaoudzi', 'GSA ISSOUFALI Jumbo Score', 'GSA ISSOUFALI Mamoudzou']
        if str(pnr.get_pnr_office()).strip() in agence_name_uniformised:
            return str(pnr.get_pnr_office()).strip().removeprefix("GSA ISSOUFALI")
        return pnr.get_pnr_office()
    except:
        return None


# ------------------------ STATISTIQUES -----------------------------------------------------------------

def graph_view(request):

    context = {}
    
    total_pnr_for_week = get_total_pnr_for_week()
    context['all_pnr_count'] = total_pnr_for_week['all_pnr_count']
    context['last_week_pnr_count'] = total_pnr_for_week['last_week_pnr_count']

    context['total_pnr'] = Pnr.objects.count()

    today_date = datetime.today().date()
    # formatted_date = today_date.strftime('%Y-%m-%d')
    formatted_date= '2023-06-22'
    context['pnr_of_today'] = get_pnr_of_today(formatted_date)
    context['pnr_invoiced_today'] = get_pnr_invoiced_today(formatted_date)
    context['pnr_to_invoice'] = get_pnr_to_invoice_today(formatted_date)

    pnr_difference = get_pnr_difference()

    context['pourcentage_pnr_remonte'] = pnr_difference['pourcentage_pnr_remonte']
    context['pourcentage_pnr_invoiced'] = pnr_difference['pourcentage_pnr_invoiced']
    context['pourcentage_pnr_to_invoice'] = pnr_difference['pourcentage_pnr_to_invoice']

    context['pnr_invoiced'] = get_pnr_invoices_by_month()
    context['pnr_created'] = get_pnr_created_by_month()
    
    
    return render(request, 'stat.html', context)  

def passenger_graph_view(request):

    context = {}
    passenger_by_age_data = get_passenger_by_age()
    context['passenger_by_age'] = passenger_by_age_data['data']
    context['total_passenger'] = passenger_by_age_data['total']

    context['all_data'] = get_destination_by_month()

    context['all_data_origin'] = get_origin_by_month()

    context['all_data_airline'] = get_stat_airlines()
    total_pnr_for_week = get_total_pnr_for_week()
    context['all_pnr_count'] = total_pnr_for_week['all_pnr_count']
    context['last_week_pnr_count'] = total_pnr_for_week['last_week_pnr_count']

    most_used_airlines = get_most_used_airlines()
    context['total_most_used_airlines'] = len(most_used_airlines)
    context['most_used_airlines'] = most_used_airlines

    context['passenger_by_month'] = get_passenger_by_month()

    context['passenger_of_today'] = get_passenger_of_today()

    context['passenger_of_the_month'] = get_passenger_of_the_month()

    
    
    return render(request, 'stat/passenger_stat.html', context)  


def anomaly_graph_view(request):

    context = {}
    context['anomaly_by_month'] =  get_anomaly_created_by_month()
    context['anomaly_by_user'] = get_anomaly_created_by_user()

    context['anomaly_of_today'] = get_anomaly_of_today()
    context['anomaly_of_this_month'] = get_anomaly_of_this_month()
    context['total_anomaly'] = get_total_anomaly()
    context['anomaly_non_traite'] = get_anomaly_non_traite()


    
    return render(request, 'stat/anomalie_stat.html', context)  


def user_graph_view(request):

    context = {}
    
    context['all_data_user'] = get_stat_user_comment()
    context['all_data_user_pnr'] = get_stat_user_pnr()

    return render(request, 'stat/user_stat.html', context)  


def get_stat_airlines():

    all_year = PnrAirSegments.objects.annotate(year=ExtractYear('departuretime')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')
    all_data_by_month = []

    for month in range(1, 13):
        all_data = {}
        all_data['month'] = month
        all_data['data'] = []
        with connection.cursor() as cursor:
            for element in all_year:
                if element['year'] is not None:
                    data = {}
                    cursor.execute("""
                    SELECT servicecarrier_id, count(*) as total
                    FROM v_pnr_passenger
                    WHERE EXTRACT(MONTH FROM departuretime) = %s
                    AND EXTRACT(YEAR FROM departuretime) = %s
                    GROUP BY servicecarrier_id order by total desc;
                """, [month,element['year']])
                    
                    # Récupérer les résultats
                    results = cursor.fetchall()

                    data['year']=(str(element['year']))
                    data['data'] = []
                    for result in results:
                        airline = Airline.objects.get(pk = result[0])
                        data['data'].append({"y":result[1],"label":airline.name}) 
                        
                    all_data['data'].append(data)
        all_data_by_month.append(all_data)
    return all_data_by_month
    
def get_stat_user_comment():
    
    all_year = PnrAirSegments.objects.annotate(year=ExtractYear('departuretime')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')
    all_data_by_month = []

    for month in range(1, 13):
        all_data = {}
        all_data['month'] = month
        all_data['data'] = []
        with connection.cursor() as cursor:
            for element in all_year:
                if element['year'] is not None:
                    data = {}
                    cursor.execute("""
                    select username, count(*) as total from v_comment_user vcu 
                    WHERE EXTRACT(MONTH FROM creation_date) = %s
                    AND EXTRACT(YEAR FROM creation_date) = %s
                    GROUP BY username order by total desc;
                """, [month,element['year']])
                    
                    # Récupérer les résultats
                    results = cursor.fetchall()

                    data['year']=(str(element['year']))
                    data['data'] = []
                    for result in results:
                        data['data'].append({"y":result[1],"label":result[0]}) 
                        
                    all_data['data'].append(data)
        all_data_by_month.append(all_data)
    return all_data_by_month

def get_stat_user_pnr():

    all_year = PnrAirSegments.objects.annotate(year=ExtractYear('departuretime')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')

    all_data_by_month = []

    for month in range(1, 13):
        all_data = {}
        all_data['month'] = month
        all_data['data'] = []
        with connection.cursor() as cursor:
            for element in all_year:
                if element['year'] is not None:
                    data = {}
                    cursor.execute("""
                    select agent_id, count(*) as total from t_pnr tp 
                    WHERE EXTRACT(MONTH FROM system_creation_date) = %s
                    AND EXTRACT(YEAR FROM system_creation_date) = %s
                    GROUP BY agent_id order by total desc;
                """, [month,element['year']])
                    
                    # Récupérer les résultats
                    results = cursor.fetchall()

                    data['year']=(str(element['year']))
                    data['data'] = []
                    for result in results:
                        if result[0] is not None:
                            user = User.objects.get(pk = result[0])
                            data['data'].append({"y":result[1],"label":user.username}) 
                        
                    all_data['data'].append(data)
        all_data_by_month.append(all_data)
    return all_data_by_month
  

def get_destination_by_month():
    
    all_year = PnrAirSegments.objects.annotate(year=ExtractYear('departuretime')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')
    all_data_by_month = []

    for month in range(1, 13):
        all_data = {}
        all_data['month'] = month
        all_data['data'] = []
        with connection.cursor() as cursor:
            for element in all_year:
                if element['year'] is not None:
                    data = {}

                    cursor.execute("""
                        SELECT ta.municipality, SUM(g.total) ::INTEGER AS total_passengers
                            FROM (
                                SELECT codedest_id, COUNT(*) AS total
                                FROM v_pnr_passenger
                                WHERE EXTRACT(MONTH FROM departuretime) = %s
                                AND EXTRACT(YEAR FROM departuretime) = %s
                                GROUP BY codedest_id
                                HAVING COUNT(*) > 20
                            ) g
                            JOIN t_airports ta ON g.codedest_id = ta.iata_code
                            GROUP BY ta.municipality
                            ORDER BY total_passengers DESC;
                    """, [month,element['year']])

                    # Récupérer les résultats
                    results = cursor.fetchall()

                    data['year']=(str(element['year']))
                    data['data'] = []
                    for result in results:
                        
                        if result[0] is None:
                            data['data'].append({"y":result[1],"label":"Inconnu"})
                        else:
                            data['data'].append({"y":result[1],"label":result[0]})

                    all_data['data'].append(data)
        all_data_by_month.append(all_data)

    return all_data_by_month

def get_origin_by_month():
    # month = datetime.datetime.now().month
    all_year = PnrAirSegments.objects.annotate(year=ExtractYear('departuretime')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')
    all_data_by_month = []

    for month in range(1, 13):
        all_data = {}
        all_data['month'] = month
        all_data['data'] = []
        with connection.cursor() as cursor:
            for element in all_year:
                if element['year'] is not None:
                    data = {}

                    cursor.execute("""
                        SELECT ta.municipality, SUM(g.total) ::INTEGER AS total_passengers
                        FROM (
                        SELECT codeorg_id, COUNT(*) AS total
                        FROM v_pnr_passenger
                        WHERE EXTRACT(MONTH FROM departuretime) = %s
                        AND EXTRACT(YEAR FROM departuretime) = %s
                        GROUP BY codeorg_id
                        HAVING COUNT(*) > 20
                        ) g
                        JOIN t_airports ta ON g.codeorg_id = ta.iata_code
                        GROUP BY ta.municipality
                        ORDER BY total_passengers DESC;

                    """, [month,element['year']])

                    # Récupérer les résultats
                    results = cursor.fetchall()

                    data['year']=(str(element['year']))
                    data['data'] = []
                    for result in results:
                        if result[0] is None:
                            data['data'].append({"y":result[1],"label":"Inconnu"})
                        else:
                            data['data'].append({"y":result[1],"label":result[0]})

                        
                    all_data['data'].append(data)
        all_data_by_month.append(all_data)
   

    return all_data_by_month

def get_passenger_by_age():


    queryset = Passenger.objects.annotate(lower_designation=Lower('designation')).values('lower_designation').annotate(total=Count('id'))
    data = []
    total_adt = 0
    total_chd = 0
    total = 0
    for item in queryset:
        total += item['total']

    for item in queryset:
        if item['lower_designation'] in ['bébé']:
            data.append({ "label": "Bébé", "y": (item['total'] * 100)/total })

        if item['lower_designation'] in ['enfant','inf','chd']:
            total_chd += item['total']

        if item['lower_designation'] in ['mrs','ms','mlle','adt','m.','mme','mr','mstr','dr','cnn','Inconnu','yth']:
            total_adt += item['total'] 
        
        
    data.append({ "label": "Adulte", "y": (total_adt * 100)/total }) 
    data.append({ "label": "Enfant", "y":  (total_chd * 100)/total }) 
    
    context = {'total':total, 'data':data}
    return context

# get all the date of the week of a specific date
def get_all_date_of_the_week(date):
    
    # Calculer le début et la fin de la semaine
    debut_semaine = date - timedelta(days=date.weekday())
    # fin_semaine = debut_semaine + timedelta(days=6)

    # Afficher les dates de tous les jours de la semaine
    dates_semaine = [debut_semaine + timedelta(days=i) for i in range(7)]

    return dates_semaine


# get the total of pnr for the week and the week before 
def get_total_pnr_for_week():
    context = {}
    aujourdhui = datetime.today().date()
    all_week_date = [date.strftime('%Y-%m-%d') for date in get_all_date_of_the_week(aujourdhui)]

    last_week = aujourdhui - timedelta(days=(aujourdhui.weekday() + 6))
    print('last_week : ', last_week)
    all_last_week_date = [date.strftime('%Y-%m-%d') for date in get_all_date_of_the_week(last_week)]


    print('this week : ', all_week_date)
    print('last week all date: ', all_last_week_date)

    all_pnr_count = Pnr.objects.filter(system_creation_date__date__in=all_week_date).count()
    last_week_pnr_count = Pnr.objects.filter(system_creation_date__date__in=all_last_week_date).count()

    context['all_pnr_count'] = all_pnr_count
    context['last_week_pnr_count'] = last_week_pnr_count

    return context

def get_passenger_by_month():
    
    all_year = PnrAirSegments.objects.annotate(year=ExtractYear('departuretime')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')
    all_data = []
    

    with connection.cursor() as cursor:
        for element in all_year:
            if element['year'] is not None:
                data = {}
                data['year']=(str(element['year']))
                data['data'] = []
                cursor.execute("""
                    WITH all_months AS (
                    SELECT generate_series(1, 12) AS mois
                    )
                    , passengers_by_month AS (
                    SELECT TO_CHAR(departuretime, 'TMMonth') AS month, COUNT(passenger_id) AS total, EXTRACT(month FROM departuretime) AS mois
                    FROM v_pnr_passenger
                    WHERE EXTRACT(YEAR FROM departuretime) = %s
                    GROUP BY TO_CHAR(departuretime, 'TMMonth'), EXTRACT(month FROM departuretime)
                    )
                    SELECT 
                    TO_CHAR(TO_DATE(m.mois::text, 'MM'), 'TMMonth') AS month, 
                    COALESCE(p.total, 0) AS total, 
                    m.mois
                    FROM 
                    all_months m
                    LEFT JOIN 
                    passengers_by_month p 
                    ON 
                    m.mois = p.mois
                    ORDER BY 
                    m.mois;

                """, [element['year']])

                # Récupérer les résultats
                results = cursor.fetchall()
                for result in results:
                    data['data'].append({"label":result[0],"y":result[1]})
                all_data.append(data)
                

    return all_data

def get_most_used_airlines():
    most_used_airlines = []
    with connection.cursor() as cursor:

        cursor.execute("""
            select servicecarrier_id , count(servicecarrier_id) from v_pnr_passenger vpp 
            group by servicecarrier_id 
            having count(servicecarrier_id) >20
        """)

        # Récupérer les résultats
        results = cursor.fetchall()

        for result in results:
            airline = Airline.objects.get(pk = result[0])
            most_used_airlines.append({"y":result[1],"label":airline.name})
    return most_used_airlines

def get_passenger_of_today():
    passenger_of_today = 0
    # now = datetime.now()
    date = '2023-06-21'
    now = datetime.strptime(date, "%Y-%m-%d")

    with connection.cursor() as cursor:

        cursor.execute("""
            select extract (day from departuretime) as day, count(passenger_id) from v_pnr_passenger vpp 
            where extract (day from departuretime) = %s
            and extract (month from departuretime) = %s
            and extract (year from departuretime) = %s
            group by day
        """, [now.day,now.month,now.year])

        # Récupérer les résultats
        results = cursor.fetchall()

        for result in results:
            passenger_of_today = result[1]

    return passenger_of_today
            
def get_passenger_of_the_month():
    passenger_of_the_month = 0
    # now = datetime.now()
    date = '2023-06-21'
    now = datetime.strptime(date, "%Y-%m-%d")

    with connection.cursor() as cursor:

        cursor.execute("""
            select extract (month from departuretime) as month, count(passenger_id) from v_pnr_passenger vpp 
            where extract (month from departuretime) = %s
            and extract (year from departuretime) = %s
            group by month
        """, [now.month,now.year])

        # Récupérer les résultats
        results = cursor.fetchall()

        for result in results:
            passenger_of_the_month = result[1]

    return passenger_of_the_month
            
def get_pnr_of_today(date):
    pnr_count_by_date=0
    

    with connection.cursor() as cursor:

        cursor.execute("""
            select date(system_creation_date) ,count(*) from t_pnr tp 
            where date(system_creation_date) = %s
            group by date(system_creation_date)
        """, [date])

        # Récupérer les résultats
        results = cursor.fetchall()

        for result in results:
            pnr_count_by_date = result[1]

    return pnr_count_by_date

def get_pnr_invoiced_today(date):
    
    pnr_count_by_date=0

    with connection.cursor() as cursor:

        cursor.execute("""
            select date(system_creation_date) ,count(*) from t_pnr tp 
            where date(system_creation_date) = %s
            and is_invoiced = true 
            group by date(system_creation_date)
        """, [date])

        # Récupérer les résultats
        results = cursor.fetchall()

        for result in results:
            pnr_count_by_date = result[1]

    return pnr_count_by_date

def get_pnr_to_invoice_today(date):
    pnr_to_invoice = get_pnr_of_today(date) - get_pnr_invoiced_today(date)

    return pnr_to_invoice

def get_pnr_difference():
    context = {}
    today_date = datetime.today().date()
    # formatted_date = today_date.strftime('%Y-%m-%d')
    formatted_date= '2023-06-22'
    yesterday_date= '2023-06-21'

    # yesterday_date = today_date - timedelta(days=1)

    pnr_of_today = get_pnr_of_today(formatted_date)
    pnr_invoiced_today = get_pnr_invoiced_today(formatted_date)
    pnr_to_invoice_today = get_pnr_to_invoice_today(formatted_date)

    pnr_of_yesterday= get_pnr_of_today(yesterday_date)
    pnr_invoiced_yesterday= get_pnr_invoiced_today(yesterday_date)
    pnr_to_invoice_yesterday= get_pnr_to_invoice_today(yesterday_date)

    pnr_remonte_diff = pnr_of_today - pnr_of_yesterday
    pourcentage_pnr_remonte = (pnr_remonte_diff * 100)/pnr_of_today

    pnr_invoiced_diff = pnr_invoiced_today - pnr_invoiced_yesterday
    pourcentage_pnr_invoiced = (pnr_invoiced_diff * 100)/pnr_invoiced_today

    pnr_to_invoice_diff = pnr_to_invoice_today - pnr_to_invoice_yesterday
    pourcentage_pnr_to_invoice = (pnr_to_invoice_diff * 100)/pnr_to_invoice_today

    context['pourcentage_pnr_remonte'] = pourcentage_pnr_remonte
    context['pourcentage_pnr_invoiced'] = pourcentage_pnr_invoiced
    context['pourcentage_pnr_to_invoice'] = pourcentage_pnr_to_invoice

    return context

def get_pnr_invoices_by_month():
    all_year = Pnr.objects.annotate(year=ExtractYear('system_creation_date')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')
    all_data = []
    

    with connection.cursor() as cursor:
        for element in all_year:
            if element['year'] is not None:
                data = {}
                data['year']=(str(element['year']))
                data['data'] = []
                cursor.execute("""
                    WITH all_months AS (
                    SELECT generate_series(1, 12) AS mois
                    )
                    , pnr_by_month AS (
                    SELECT TO_CHAR(system_creation_date, 'TMMonth') AS month, COUNT(*) AS total, EXTRACT(month FROM system_creation_date) AS mois
                    FROM t_pnr
                    WHERE is_invoiced = true 
                    AND EXTRACT(year FROM system_creation_date) = %s
                    GROUP BY TO_CHAR(system_creation_date, 'TMMonth'), EXTRACT(month FROM system_creation_date)
                    )
                    SELECT 
                    TO_CHAR(TO_DATE(m.mois::text, 'MM'), 'TMMonth') AS month, 
                    COALESCE(p.total, 0) AS total, 
                    m.mois
                    FROM 
                    all_months m
                    LEFT JOIN 
                    pnr_by_month p 
                    ON 
                    m.mois = p.mois
                    ORDER BY 
                    m.mois;

                """, [element['year']])

                # Récupérer les résultats
                results = cursor.fetchall()
                for result in results:
                    data['data'].append({"label":result[0],"y":result[1]})
                all_data.append(data)
                

    return all_data


def get_pnr_created_by_month():
    all_year = Pnr.objects.annotate(year=ExtractYear('system_creation_date')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')
    all_data = []
    

    with connection.cursor() as cursor:
        for element in all_year:
            if element['year'] is not None:
                data = {}
                data['year']=(str(element['year']))
                data['data'] = []
                cursor.execute("""
                        WITH all_months AS (
                            SELECT generate_series(1, 12) AS mois
                        )
                        , pnr_by_month AS (
                            SELECT TO_CHAR(system_creation_date, 'TMMonth') AS month, COUNT(*) AS total, EXTRACT(month FROM system_creation_date) AS mois
                            FROM t_pnr
                            WHERE EXTRACT(year FROM system_creation_date) = %s
                            GROUP BY TO_CHAR(system_creation_date, 'TMMonth'), EXTRACT(month FROM system_creation_date)
                        )
                        SELECT 
                            TO_CHAR(TO_DATE(m.mois::text, 'MM'), 'TMMonth') AS month, 
                            COALESCE(p.total, 0) AS total, 
                            m.mois
                        FROM 
                            all_months m
                        LEFT JOIN 
                            pnr_by_month p 
                        ON 
                            m.mois = p.mois
                        ORDER BY 
                            m.mois;

                """, [element['year']])

                # Récupérer les résultats
                results = cursor.fetchall()
                for result in results:
                    data['data'].append({"label":result[0],"y":result[1]})
                all_data.append(data)
            
    return all_data

def get_anomaly_created_by_month():
    all_year = Comment.objects.annotate(year=ExtractYear('creation_date')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')
    all_data = []
    

    with connection.cursor() as cursor:
        for element in all_year:
            if element['year'] is not None:
                data = {}
                data['year']=(str(element['year']))
                data['data'] = []

                cursor.execute("""
                WITH all_months AS (
                    SELECT generate_series(1, 12) AS month_num
                )
                , comments_by_month AS (
                    SELECT EXTRACT(month FROM creation_date) AS mois, COUNT(*) AS total
                    FROM t_comment
                    WHERE EXTRACT(YEAR FROM creation_date) = %s
                    GROUP BY EXTRACT(month FROM creation_date)
                )
                SELECT 
                    TO_CHAR(TO_DATE(m.month_num::text, 'MM'), 'TMMonth') AS month, 
                    COALESCE(c.total, 0) AS total, 
                    m.month_num AS mois
                FROM 
                    all_months m
                LEFT JOIN 
                    comments_by_month c 
                ON 
                    m.month_num = c.mois
                ORDER BY 
                    m.month_num;
                """, [element['year']])

                # Récupérer les résultats
                results = cursor.fetchall()
                for result in results:
                    data['data'].append({"label":result[0],"y":result[1]})
                all_data.append(data)
            
    return all_data

def get_anomaly_created_by_user():
    all_year = Comment.objects.annotate(year=ExtractYear('creation_date')).values('year').distinct().filter(year__gte=datetime(2023, 1, 1).year).order_by('year')
    all_data_by_month = []

    for month in range(1, 13):
        all_data = {}
        all_data['month'] = month
        all_data['data'] = []
        with connection.cursor() as cursor:
            for element in all_year:
                if element['year'] is not None:
                    data = {}

                    cursor.execute("""
                        select tc.user_id_id ,count(*) as total from t_comment tc
                        WHERE EXTRACT(MONTH FROM creation_date) = %s
                        AND EXTRACT(YEAR FROM creation_date) = %s
                        group by user_id_id;
                    """, [month,element['year']])

                    # Récupérer les résultats
                    results = cursor.fetchall()

                    data['year']=(str(element['year']))
                    data['data'] = []
                    for results in results:
                        user = User.objects.get(pk = results[0])
                        data['data'].append({"y":results[1],"label":user.username}) 

                    all_data['data'].append(data)
        all_data_by_month.append(all_data)

    return all_data_by_month

def get_anomaly_of_today():
    
    anomaly_count_by_date=0
    # now = datetime.now()
    date = '2023-06-21'
    now = datetime.strptime(date, "%Y-%m-%d")

    with connection.cursor() as cursor:

        cursor.execute("""
            select date(creation_date) ,count(*) from t_comment 
            where date(creation_date) = %s
            group by date(creation_date)
        """, [date])

        # Récupérer les résultats
        results = cursor.fetchall()

        for result in results:
            anomaly_count_by_date = result[1]

    return anomaly_count_by_date

def get_anomaly_of_this_month():
    
    anomaly_count_by_date=0
    # now = datetime.now()
    date = '2023-06-21'
    now = datetime.strptime(date, "%Y-%m-%d")

    with connection.cursor() as cursor:

        cursor.execute("""
            select extract (month from creation_date) as month, count(*) from t_comment 
            where extract (month from creation_date) = %s
            and extract (year from creation_date) = %s
            group by month
        """, [now.month,now.year])

        # Récupérer les résultats
        results = cursor.fetchall()

        for result in results:
            anomaly_count_by_date = result[1]

    return anomaly_count_by_date

def get_total_anomaly():
    
    anomaly_count=0

    with connection.cursor() as cursor:

        cursor.execute("""
            select count(*) from t_comment 
        """)

        # Récupérer les résultats
        results = cursor.fetchall()

        for result in results:
            anomaly_count = result[0]

    return anomaly_count

def get_anomaly_non_traite():
    
    anomaly_count=0

    with connection.cursor() as cursor:

        cursor.execute("""
            select count(*) from t_comment 
            where state = false
        """)

        # Récupérer les résultats
        results = cursor.fetchall()

        for result in results:
            anomaly_count = result[0]

    return anomaly_count












































